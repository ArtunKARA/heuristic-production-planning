#!/usr/bin/env python3
"""Build the single, self-contained Excel report for the paper tables.

Run with:
    C:\\Users\\Artun\\anaconda3\\python.exe Doc\\build_tables_xlsx.py

The script does NOT execute any heuristic.  It only consumes CSV / manifest
artifacts already on disk and writes one workbook:

    Doc/benchmark_outputs/tables_report.xlsx

Auto-discovery
==============

It scans Doc/benchmark_outputs/ for sub-directories that look like a
benchmark run (must contain manifest.json and a csv/ folder).  Every
qualifying run is loaded and merged.  When more methods become available
later (for example GA / Tabu variants in a new paper_comparison_main_3
folder), just drop the run next to the existing one and re-run this
script; the workbook regenerates with full coverage and the "missing
methods" placeholder shrinks automatically.

Notation kept consistent everywhere in the workbook (per reviewer feedback)
========================================================================

* NP        = population size  (NP = 20 in the current run)
* n_iter    = outer iteration count {10, 20, 30, 50, 100, 200}
* FE        = evaluator (fitness) call count = eval_calls_total
* metric    = total_score = total_cost + hard_total ; lower is better
* problem-size is described by instance features (products / machines /
  molds / lots = order lines / weekly buckets / shifts), never by n_iter
"""

from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
BENCH_ROOT = ROOT / "Doc" / "benchmark_outputs"
INPUT_JSON = ROOT / "Doc" / "SampleData" / "example_input.json"
OUT_XLSX = BENCH_ROOT / "tables_report.xlsx"

NP_VALUE = 20

ALL_METHODS_CANONICAL = [
    "ga", "tabu", "gatabu", "ga_tabu_inline", "ga_tabu_topk",
    "hho", "hmpa", "cssrank",
]
N_ITER_CANONICAL = [10, 20, 30, 50, 100, 200]


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, color="1F4E78", size=12)
TITLE_FILL = PatternFill("solid", fgColor="305496")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
PLACEHOLDER_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FONT = Font(italic=True, color="595959")
WARN_FONT = Font(italic=True, bold=True, color="9C0006")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Auto-discovery + CSV loading
# ---------------------------------------------------------------------------

def discover_runs(root: Path) -> List[Path]:
    """Return benchmark run directories under root, sorted by completed_at."""
    if not root.exists():
        return []
    candidates: List[Tuple[datetime, Path]] = []
    for child in sorted(root.iterdir()):
        if not child.is_dir():
            continue
        manifest = child / "manifest.json"
        csv_dir = child / "csv"
        if not (manifest.exists() and csv_dir.exists()):
            continue
        try:
            m = json.loads(manifest.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        when = m.get("completed_at") or m.get("created_at") or ""
        try:
            ts = datetime.fromisoformat(when) if when else datetime.min
        except ValueError:
            ts = datetime.min
        candidates.append((ts, child))
    candidates.sort()
    return [p for _, p in candidates]


def first_csv(csv_dir: Path, *patterns: str) -> Path | None:
    """Return the newest CSV that matches any of the glob patterns."""
    hits: List[Tuple[float, Path]] = []
    for pat in patterns:
        for p in csv_dir.glob(pat):
            if p.is_file() and p.stat().st_size > 0:
                hits.append((p.stat().st_mtime, p))
    if not hits:
        return None
    hits.sort()
    return hits[-1][1]


def read_csv(path: Path | None) -> List[Dict[str, str]]:
    if path is None:
        return []
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def load_run(run_dir: Path) -> Dict[str, Any]:
    """Load all standard artifacts for a single benchmark run directory."""
    manifest = json.loads((run_dir / "manifest.json").read_text(encoding="utf-8"))
    csv_dir = run_dir / "csv"
    return {
        "dir": run_dir,
        "manifest": manifest,
        "summary_n": read_csv(first_csv(csv_dir,
            "summary_by_method_n*.csv")),
        "summary_all": read_csv(first_csv(csv_dir,
            "summary_by_method.csv", "summary_by_method[0-9]*.csv")),
        "fairness_n": read_csv(first_csv(csv_dir,
            "fairness_by_method_n*.csv")),
        "fairness_all": read_csv(first_csv(csv_dir,
            "fairness_by_method.csv", "fairness_by_method[0-9]*.csv")),
        "fe_budget": read_csv(first_csv(csv_dir,
            "fe_budget_summary*.csv", "fe_budget_at_*.csv")),
        "convergence_path": first_csv(csv_dir,
            "convergence_history*.csv"),
    }


def merge_runs(runs: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Merge multiple runs: later runs override earlier ones per (method, n_iter)."""
    merged: Dict[str, List[Dict[str, Any]]] = {
        "summary_n": [],
        "summary_all": [],
        "fairness_n": [],
        "fairness_all": [],
        "fe_budget": [],
    }

    def upsert(key_fn, table_key: str, run: Dict[str, Any]) -> None:
        existing = {key_fn(r): i for i, r in enumerate(merged[table_key])}
        for row in run[table_key]:
            k = key_fn(row)
            if k in existing:
                merged[table_key][existing[k]] = row
            else:
                merged[table_key].append(row)
                existing[k] = len(merged[table_key]) - 1

    for run in runs:
        upsert(lambda r: (r.get("method"), r.get("n_iter")),
               "summary_n", run)
        upsert(lambda r: r.get("method"),
               "summary_all", run)
        upsert(lambda r: (r.get("method"), r.get("n_iter")),
               "fairness_n", run)
        upsert(lambda r: r.get("method"),
               "fairness_all", run)
        upsert(lambda r: (r.get("method"), r.get("n_iter"), r.get("fe_budget")),
               "fe_budget", run)
    return merged


def methods_present(merged: Dict[str, Any]) -> List[str]:
    return sorted({r["method"] for r in merged["summary_n"] if r.get("method")},
                  key=lambda m: ALL_METHODS_CANONICAL.index(m)
                  if m in ALL_METHODS_CANONICAL else 99)


def methods_missing(present: Sequence[str]) -> List[str]:
    s = set(present)
    return [m for m in ALL_METHODS_CANONICAL if m not in s]


# ---------------------------------------------------------------------------
# Workbook utilities
# ---------------------------------------------------------------------------

def style_header(ws, row_idx: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_table(ws, start_row: int, headers: Sequence[str],
                rows: Iterable[Sequence[Any]],
                number_formats: Dict[int, str] | None = None) -> int:
    number_formats = number_formats or {}
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)
    style_header(ws, start_row, len(headers))

    r = start_row + 1
    for row in rows:
        is_placeholder = any(v == "-" for v in row[3:])
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = BORDER
            fmt = number_formats.get(c - 1)
            if fmt is not None and value not in ("-", None):
                cell.number_format = fmt
            if isinstance(value, str):
                cell.alignment = WRAP
            if is_placeholder:
                cell.fill = PLACEHOLDER_FILL
        r += 1
    return r


def big_title(ws, row_idx: int, text: str, span: int = 8) -> int:
    ws.cell(row=row_idx, column=1, value=text)
    cell = ws.cell(row=row_idx, column=1)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=span)
    ws.row_dimensions[row_idx].height = 26
    return row_idx + 1


def section_title(ws, row_idx: int, text: str, span: int = 8) -> int:
    ws.cell(row=row_idx, column=1, value=text)
    cell = ws.cell(row=row_idx, column=1)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=span)
    return row_idx + 1


def add_note(ws, row_idx: int, text: str, span: int = 8, warn: bool = False) -> int:
    ws.cell(row=row_idx, column=1, value=text)
    cell = ws.cell(row=row_idx, column=1)
    cell.font = WARN_FONT if warn else NOTE_FONT
    cell.alignment = WRAP
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=span)
    approx = max(1, len(text) // 95 + text.count("\n"))
    ws.row_dimensions[row_idx].height = max(18, approx * 15)
    return row_idx + 1


def autosize_columns(ws, min_width: int = 10, max_width: int = 70) -> None:
    widths: Dict[int, int] = defaultdict(lambda: min_width)
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None or not isinstance(cell.column, int):
                continue
            text = str(cell.value)
            for line in text.splitlines():
                widths[cell.column] = min(max_width, max(widths[cell.column], len(line) + 2))
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def fnum(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_scope(wb: Workbook, present: List[str], missing: List[str]) -> None:
    ws = wb.create_sheet("0. Scope & terminology")
    is_complete = not missing

    title = (
        "All-method comparison tables (NP = 20)"
        if is_complete else
        f"NP=20 re-evaluation tables  |  methods present: {', '.join(present)}"
    )
    r = 1
    r = big_title(ws, r, title + "  |  instance: DB_SAMPLE", span=4)
    r += 1

    r = section_title(ws, r, "Scope of this workbook", span=4)
    if is_complete:
        r = add_note(ws, r,
            "All 8 methods are present.  Sheets 3-6 are the main all-method comparison "
            "tables; sheet 9 lists per-method run provenance.",
            span=4)
    else:
        r = add_note(ws, r,
            "Tables in sheets 3-6 cover ONLY the methods listed above.  They are a "
            "NP=20 re-evaluation of those methods, NOT the main all-method comparison. "
            "Missing methods: " + ", ".join(missing) + ".  See sheet 9 (placeholder).",
            span=4, warn=True)
    r += 1

    r = section_title(ws, r, "Notation used everywhere", span=4)
    notation = [
        ("NP", "Population size.  NP = 20 in the current data "
               "(hho_hawks = hmpa_predators = css_particles = 20)."),
        ("n_iter", "Outer iteration count {10, 20, 30, 50, 100, 200}.  A search-budget "
                   "parameter, NOT a problem-size parameter."),
        ("FE", "Evaluator (fitness) call count = eval_calls_total per run.  "
               "Fair-budget cut points reported at FE in {500, 1000, max-FE}."),
        ("metric", "total_score = total_cost + hard_total.  Lower is better.  "
                   "Relation verified row-by-row in run_results.csv."),
        ("Problem size", "Described by instance features (next sheet): products / "
                         "machines / molds / order lines (= lots) / weekly buckets / "
                         "shifts.  Avoid 'N = 200' for the iteration count; "
                         "write 'n_iter = 200' instead."),
    ]
    r = write_table(ws, r, ["Term", "Definition"], notation)
    r += 1

    r = section_title(ws, r, "Per-method NP setting in the current data", span=4)
    method_params = [
        ("hho", "hho_hawks = 20", "time_shift_max=8h, bucket_shift_max=2, qty_jitter_max=20%, "
                                  "machine_swap_max=50%, mold_swap_max=50%"),
        ("hmpa", "hmpa_predators = 20", "local_trials=3, local_search_every=4, "
                                        "local_radius=0.08, fads_rate=0.15"),
        ("cssrank", "css_particles = 20", "top_ratio=0.4, damping=0.72, accel=1.25, "
                                          "elite_pull=0.28, reset_rate=0.12, noise_scale=0.03"),
    ]
    r = write_table(ws, r,
                    ["method", "population control = NP", "method-specific controls"],
                    [row for row in method_params if row[0] in present] or method_params)
    autosize_columns(ws)


def build_sources(wb: Workbook, runs: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("1. Sources & overview")
    r = 1
    r = big_title(ws, r, "Benchmark run sources merged into this workbook", span=6)
    r += 1

    headers = ["run directory", "frame_id", "methods", "n_iter", "runs_per_n",
               "completed_at"]
    rows = []
    for run in runs:
        m = run["manifest"]
        rows.append([
            str(run["dir"].relative_to(ROOT)),
            m.get("frame_id"),
            ", ".join(m.get("methods", [])),
            ", ".join(str(x) for x in m.get("iterations", [])),
            m.get("runs_per_n"),
            m.get("completed_at"),
        ])
    r = write_table(ws, r, headers, rows)
    r += 1

    r = section_title(ws, r, "Aggregate counts", span=6)
    total_jobs = sum(int(run["manifest"].get("completed_jobs") or 0) for run in runs)
    elapsed_sum = sum(float(run["manifest"].get("elapsed_sec") or 0) for run in runs)
    methods_union = sorted({m for run in runs for m in run["manifest"].get("methods", [])},
                           key=lambda m: ALL_METHODS_CANONICAL.index(m)
                           if m in ALL_METHODS_CANONICAL else 99)
    summary = [
        ("Run directories merged", len(runs)),
        ("Methods covered (union)", ", ".join(methods_union)),
        ("Total completed jobs (sum)", total_jobs),
        ("Total wall-clock (sec, sum)", round(elapsed_sum, 2)),
        ("Total wall-clock (hh:mm:ss)",
         f"{int(elapsed_sum // 3600):02d}:{int((elapsed_sum % 3600) // 60):02d}:"
         f"{int(elapsed_sum % 60):02d}"),
    ]
    r = write_table(ws, r, ["Field", "Value"], summary)
    autosize_columns(ws)


def build_instance(wb: Workbook, problem_data: Dict[str, Any]) -> None:
    ws = wb.create_sheet("2. Instance description")

    pd_ = problem_data
    buckets = pd_["time_buckets"]
    products = pd_["products"]
    machines = pd_["resources"]["machines"]
    molds = pd_["resources"]["molds"]
    compat = pd_["compatibility"]

    order_entries: List[Dict[str, Any]] = []
    for wrap in pd_["orders"]:
        for o in wrap["orders"]:
            order_entries.append({
                "product": wrap["product_code"],
                "bucket": o["time_bucket_id"],
                "qty": float(o["qty"]),
            })
    total_demand = sum(e["qty"] for e in order_entries)
    unique_demand_products = sorted({e["product"] for e in order_entries})
    unique_demand_buckets = sorted({e["bucket"] for e in order_entries})

    used_procs = sorted({s["process_code"] for p in products for s in p["process_data"]})
    total_steps = sum(len(p["process_data"]) for p in products)

    mach_by_proc: Dict[str, int] = defaultdict(int)
    for m in machines:
        mach_by_proc[m["process_code"]] += 1
    mold_by_proc: Dict[str, int] = defaultdict(int)
    for m in molds:
        mold_by_proc[m["process_code"]] += 1
    total_eye = sum(int(m.get("eye") or 0) for m in molds)

    starts = sorted(b["start_date"] for b in buckets)
    ends = sorted(b["end_date"] for b in buckets)
    work_cal = pd_.get("work_calendar", [])
    holidays = sum(1 for d in work_cal if d.get("holiday"))

    rows = [
        ("Instance code", pd_["problem_meta"]["problem_code"]),
        ("Horizon (calendar)", f"{starts[0]} -> {ends[-1]}"),
        ("Weekly time buckets (haftalik bucket)", len(buckets)),
        ("Work-calendar days", len(work_cal)),
        ("Holidays in calendar", holidays),
        ("Shift templates (vardiya)",
         f"{len(pd_['shift_templates'])} (base = {pd_['problem_meta']['base_shift_templates_code']})"),
        ("Products / SKU (urun)", len(products)),
        ("Products with demand", len(unique_demand_products)),
        ("Order lines = lots (siparis satiri = lot)", len(order_entries)),
        ("Total demand quantity (Adet)", total_demand),
        ("Demand buckets referenced", len(unique_demand_buckets)),
        ("Materials", len(pd_["materials"])),
        ("Stock entries (baslangic envanteri)", len(pd_["stocks"])),
        ("Process catalog size", len(pd_["processes"])),
        ("Distinct process codes used by products", len(used_procs)),
        ("Total product x process steps", total_steps),
        ("Machines (makine, total)", len(machines)),
        ("Machines per process",
         ", ".join(f"{p}: {mach_by_proc[p]}" for p in sorted(mach_by_proc))),
        ("Molds (kalip, total)", len(molds)),
        ("Molds per process",
         ", ".join(f"{p}: {mold_by_proc[p]}" for p in sorted(mold_by_proc))),
        ("Total mold eye count (kalip goz toplami)", total_eye),
        ("Machine-mold compatibility pairs", len(compat["machine_mold_pairs"])),
        ("Product-mold compatibility entries", len(compat["product_molds"])),
        ("Bottleneck process", "AP300 (Vulkanizasyon)"),
    ]

    r = 1
    r = big_title(ws, r,
        "Instance description of DB_SAMPLE  (problem-size = instance features)",
        span=2)
    r += 1
    r = write_table(ws, r, ["Field", "Value"], rows, number_formats={1: "#,##0"})
    r += 1
    r = add_note(ws, r,
        "Source: Doc/SampleData/example_input.json.  Lot = planning-level order line "
        "(per product x weekly bucket).  Bottleneck identified by largest pool and the "
        "only mold-bearing process: AP300 (37 machines + 20 molds + 124 compatibility pairs).",
        span=2)
    r += 1

    r = section_title(ws, r, "Demand per weekly bucket (hafta)", span=2)
    by_bucket: Dict[str, float] = defaultdict(float)
    for e in order_entries:
        by_bucket[e["bucket"]] += e["qty"]
    bucket_order = [b["id"] for b in sorted(buckets, key=lambda x: x["index"])]
    r = write_table(ws, r, ["Bucket id", "Demand qty (Adet)"],
                    [(b, by_bucket.get(b, 0.0)) for b in bucket_order],
                    number_formats={1: "#,##0"})
    r += 1
    r = section_title(ws, r, "Process steps per product", span=2)
    r = write_table(ws, r, ["Product", "Process steps"],
                    [(p["code"], len(p["process_data"])) for p in products],
                    number_formats={1: "0"})
    autosize_columns(ws)


def build_fixed_iter(wb: Workbook, merged: Dict[str, Any],
                     present: List[str], missing: List[str]) -> None:
    is_complete = not missing
    sheet_name = "3. Fixed-iteration" if is_complete else "3. NP=20 Fixed-iteration"
    title = ("Fixed-iteration result table  (NP = 20  ;  all methods)"
             if is_complete else
             f"Fixed-iteration result table  (NP = 20  ;  {', '.join(present)})")

    ws = wb.create_sheet(sheet_name)
    r = 1
    r = big_title(ws, r, title, span=11)
    r += 1
    if missing:
        r = add_note(ws, r,
            "Rows for " + ", ".join(missing) + " are placeholders (-).  See sheet 9.",
            span=11, warn=True)
        r += 1

    summary_n_by_key = {(r["method"], int(r["n_iter"])): r for r in merged["summary_n"]}

    r = section_title(ws, r,
        "Per (method, n_iter), 30 runs each, NP = 20", span=11)
    headers = ["method", "NP", "n_iter", "metric", "runs_total", "runs_ok",
               "runs_failed", "best", "mean", "worst", "std_dev"]
    rows: List[List[Any]] = []
    for method in ALL_METHODS_CANONICAL:
        for n in N_ITER_CANONICAL:
            x = summary_n_by_key.get((method, n))
            if x is None:
                rows.append([method, NP_VALUE, n, "total_score",
                             "-", "-", "-", "-", "-", "-", "-"])
            else:
                rows.append([
                    method, NP_VALUE, n, x.get("metric", "total_score"),
                    int(x["runs_total"]), int(x["runs_ok"]), int(x["runs_failed"]),
                    fnum(x["best"]), fnum(x["mean"]),
                    fnum(x["worst"]), fnum(x["std_dev"]),
                ])
    num_fmt = "#,##0.00"
    r = write_table(ws, r, headers, rows, number_formats={
        7: num_fmt, 8: num_fmt, 9: num_fmt, 10: num_fmt,
    })
    r += 1

    summary_all_by_method = {r["method"]: r for r in merged["summary_all"]}
    r = section_title(ws, r,
        "Overall per method (pooled across n_iter, 180 runs each)", span=11)
    headers2 = ["method", "NP", "metric", "runs_ok", "runs_failed",
                "best", "mean", "worst", "std_dev"]
    rows2: List[List[Any]] = []
    for method in ALL_METHODS_CANONICAL:
        x = summary_all_by_method.get(method)
        if x is None:
            rows2.append([method, NP_VALUE, "total_score",
                          "-", "-", "-", "-", "-", "-"])
        else:
            rows2.append([
                method, NP_VALUE, x.get("metric", "total_score"),
                int(x["runs_ok"]), int(x["runs_failed"]),
                fnum(x["best"]), fnum(x["mean"]),
                fnum(x["worst"]), fnum(x["std_dev"]),
            ])
    r = write_table(ws, r, headers2, rows2, number_formats={
        5: num_fmt, 6: num_fmt, 7: num_fmt, 8: num_fmt,
    })
    r += 1
    r = add_note(ws, r,
        "Source: summary_by_method_n.csv + summary_by_method.csv across the run "
        "directories listed in sheet 1.  metric = total_score; lower is better.",
        span=11)
    autosize_columns(ws)


def build_fe_budget(wb: Workbook, merged: Dict[str, Any],
                    present: List[str], missing: List[str]) -> None:
    is_complete = not missing
    sheet_name = "4. FE-budget" if is_complete else "4. NP=20 FE-budget"
    ws = wb.create_sheet(sheet_name)
    r = 1
    r = big_title(ws, r,
        ("FE-budget fair comparison  (NP = 20  ;  all methods)" if is_complete else
         f"FE-budget fair comparison  (NP = 20  ;  {', '.join(present)})"),
        span=12)
    r += 1

    parsed = []
    for x in merged["fe_budget"]:
        parsed.append({
            "method": x["method"],
            "n_iter": int(x["n_iter"]),
            "fe_budget": int(x["fe_budget"]),
            "runs_reached_budget": int(x["runs_reached_budget"]),
            "mean_score": fnum(x["mean_best_so_far_total_score"]),
            "std_score": fnum(x["std_best_so_far_total_score"]),
            "mean_cost": fnum(x["mean_best_so_far_total_cost"]),
            "std_cost": fnum(x["std_best_so_far_total_cost"]),
            "mean_hard": fnum(x["mean_best_so_far_hard_total"]),
            "median_hard": fnum(x["median_best_so_far_hard_total"]),
            "feasible_rate": fnum(x["feasible_rate_at_budget"]),
        })

    headers = ["method", "NP", "n_iter", "FE", "runs_reached_budget",
               "mean_best_so_far_total_score", "std_best_so_far_total_score",
               "mean_best_so_far_total_cost", "std_best_so_far_total_cost",
               "mean_best_so_far_hard_total", "median_best_so_far_hard_total",
               "feasible_rate_at_budget"]

    def block(start_row: int, title: str, predicate) -> int:
        r = section_title(ws, start_row, title, span=12)
        block_rows = []
        for x in parsed:
            if not predicate(x):
                continue
            block_rows.append([
                x["method"], NP_VALUE, x["n_iter"], x["fe_budget"], x["runs_reached_budget"],
                x["mean_score"], x["std_score"],
                x["mean_cost"], x["std_cost"],
                x["mean_hard"], x["median_hard"],
                x["feasible_rate"],
            ])
        block_rows.sort(key=lambda row: (row[2], row[3], row[0]))
        num = "#,##0.00"
        fmts = {i: num for i in range(5, 11)}
        fmts[11] = "0.00%"
        return write_table(ws, r, headers, block_rows, number_formats=fmts)

    r = block(r, "FE = 500   (common to n_iter in {50, 100, 200})",
              lambda x: x["fe_budget"] == 500)
    r += 1
    r = block(r, "FE = 1000  (common to n_iter in {100, 200}  ;  canonical fair budget)",
              lambda x: x["fe_budget"] == 1000)
    r += 1
    r = block(r, "FE = 3271  (n_iter = 200 natural budget)",
              lambda x: x["n_iter"] == 200 and x["fe_budget"] == 3271)
    r += 1

    r = section_title(ws, r, "Full FE-budget summary (every available FE checkpoint)", span=12)
    all_rows = []
    for x in parsed:
        all_rows.append([
            x["method"], NP_VALUE, x["n_iter"], x["fe_budget"], x["runs_reached_budget"],
            x["mean_score"], x["std_score"],
            x["mean_cost"], x["std_cost"],
            x["mean_hard"], x["median_hard"],
            x["feasible_rate"],
        ])
    all_rows.sort(key=lambda row: (row[2], row[3], row[0]))
    num = "#,##0.00"
    fmts = {i: num for i in range(5, 11)}
    fmts[11] = "0.00%"
    r = write_table(ws, r, headers, all_rows, number_formats=fmts)
    r += 1
    r = add_note(ws, r,
        "Source: fe_budget_summary*.csv across the runs in sheet 1.  Lower "
        "mean_best_so_far_total_score is better.",
        span=12)
    autosize_columns(ws)


def build_runtime(wb: Workbook, merged: Dict[str, Any],
                  present: List[str], missing: List[str]) -> None:
    is_complete = not missing
    sheet_name = "5. Runtime FE" if is_complete else "5. NP=20 Runtime FE"
    ws = wb.create_sheet(sheet_name)

    rows = []
    for x in merged["fairness_n"]:
        ec = fnum(x["mean_eval_calls_total"]) or 0.0
        es = fnum(x["mean_elapsed_sec"]) or 0.0
        eps = ec / es if es else None
        rows.append([
            x["method"], NP_VALUE, int(x["n_iter"]), int(x["runs_ok"]),
            ec, fnum(x["std_eval_calls_total"]),
            es, fnum(x["std_elapsed_sec"]),
            eps,
        ])

    by_key = {(r[0], r[2]): r for r in rows}
    rows = []
    for method in ALL_METHODS_CANONICAL:
        for n in N_ITER_CANONICAL:
            if (method, n) in by_key:
                rows.append(by_key[(method, n)])
            else:
                rows.append([method, NP_VALUE, n, "-", "-", "-", "-", "-", "-"])

    r = 1
    r = big_title(ws, r,
        ("Runtime / FE efficiency  (NP = 20  ;  all methods)" if is_complete else
         f"Runtime / FE efficiency  (NP = 20  ;  {', '.join(present)})"),
        span=9)
    r += 1
    r = section_title(ws, r,
        "Per (method, n_iter), 30 runs each, NP = 20", span=9)
    headers = ["method", "NP", "n_iter", "runs_ok",
               "mean_FE (eval_calls_total)", "std_FE",
               "mean_elapsed_sec", "std_elapsed_sec",
               "FE_per_sec (derived)"]
    num = "#,##0.00"
    fmts = {4: num, 5: num, 6: num, 7: num, 8: num}
    r = write_table(ws, r, headers, rows, number_formats=fmts)
    r += 1

    r = section_title(ws, r,
        "Overall per method (pooled across n_iter, 180 runs each)", span=9)
    by_method = {r["method"]: r for r in merged["fairness_all"]}
    rows2 = []
    for method in ALL_METHODS_CANONICAL:
        x = by_method.get(method)
        if x is None:
            rows2.append([method, NP_VALUE, "-", "-", "-", "-", "-", "-"])
        else:
            ec = fnum(x["mean_eval_calls_total"]) or 0.0
            es = fnum(x["mean_elapsed_sec"]) or 0.0
            eps = ec / es if es else None
            rows2.append([
                method, NP_VALUE, int(x["runs_ok"]),
                ec, fnum(x["std_eval_calls_total"]),
                es, fnum(x["std_elapsed_sec"]),
                eps,
            ])
    r = write_table(ws, r,
        ["method", "NP", "runs_ok", "mean_FE", "std_FE",
         "mean_elapsed_sec", "std_elapsed_sec", "FE_per_sec (derived)"],
        rows2, number_formats={3: num, 4: num, 5: num, 6: num, 7: num})
    r += 1
    r = add_note(ws, r,
        "Source: fairness_by_method_n*.csv + fairness_by_method*.csv.  "
        "FE_per_sec = mean_FE / mean_elapsed_sec (derived).",
        span=9)
    autosize_columns(ws)


def build_hard_total(wb: Workbook, merged: Dict[str, Any],
                     present: List[str], missing: List[str]) -> None:
    is_complete = not missing
    sheet_name = "6. Hard-total" if is_complete else "6. NP=20 Hard-total"
    ws = wb.create_sheet(sheet_name)

    summary_n_by_key = {(r["method"], int(r["n_iter"])): fnum(r["mean"]) or 0.0
                        for r in merged["summary_n"]}
    fair_n_by_key = {(r["method"], int(r["n_iter"])): r for r in merged["fairness_n"]}

    rows: List[List[Any]] = []
    for method in ALL_METHODS_CANONICAL:
        for n in N_ITER_CANONICAL:
            x = fair_n_by_key.get((method, n))
            if x is None:
                rows.append([method, NP_VALUE, n, "-", "-", "-", "-", "-", "-", "-"])
                continue
            score = summary_n_by_key.get((method, n), 0.0)
            cost = fnum(x["mean_best_total_cost"]) or 0.0
            median_hard = fnum(x["median_best_hard_total"]) or 0.0
            hard_share = (median_hard / score) if score else None
            cost_share = (cost / score) if score else None
            rows.append([
                method, NP_VALUE, n, int(x["runs_ok"]),
                fnum(x["feasible_rate"]),
                score, cost, median_hard, hard_share, cost_share,
            ])

    r = 1
    r = big_title(ws, r,
        ("Hard-total component table  (NP = 20  ;  all methods)" if is_complete else
         f"Hard-total component table  (NP = 20  ;  {', '.join(present)})"),
        span=10)
    r += 1
    r = section_title(ws, r,
        "Per (method, n_iter), 30 runs each, NP = 20", span=10)
    headers = ["method", "NP", "n_iter", "runs_ok", "feasible_rate",
               "mean_total_score", "mean_best_total_cost",
               "median_best_hard_total", "hard / score", "cost / score"]
    num = "#,##0.00"
    fmts = {5: num, 6: num, 7: num, 8: "0.00%", 9: "0.00%", 4: "0.00%"}
    r = write_table(ws, r, headers, rows, number_formats=fmts)
    r += 1

    r = section_title(ws, r,
        "Overall per method (pooled across n_iter, 180 runs each)", span=10)
    summary_all_score = {r["method"]: fnum(r["mean"]) or 0.0
                         for r in merged["summary_all"]}
    fair_all_by_method = {r["method"]: r for r in merged["fairness_all"]}

    rows2: List[List[Any]] = []
    for method in ALL_METHODS_CANONICAL:
        x = fair_all_by_method.get(method)
        if x is None:
            rows2.append([method, NP_VALUE, "-", "-", "-", "-", "-", "-", "-"])
            continue
        score = summary_all_score.get(method, 0.0)
        cost = fnum(x["mean_best_total_cost"]) or 0.0
        median_hard = fnum(x["median_best_hard_total"]) or 0.0
        hard_share = (median_hard / score) if score else None
        cost_share = (cost / score) if score else None
        rows2.append([
            method, NP_VALUE, int(x["runs_ok"]), fnum(x["feasible_rate"]),
            score, cost, median_hard, hard_share, cost_share,
        ])
    fmts2 = {4: num, 5: num, 6: num, 7: "0.00%", 8: "0.00%", 3: "0.00%"}
    r = write_table(ws, r,
        ["method", "NP", "runs_ok", "feasible_rate",
         "mean_total_score", "mean_best_total_cost",
         "median_best_hard_total", "hard / score", "cost / score"],
        rows2, number_formats=fmts2)
    r += 1
    r = add_note(ws, r,
        "Source: fairness_by_method_n*.csv + summary_by_method_n.csv.  "
        "total_score = total_cost + hard_total.  feasible_rate = 0 means no run "
        "reached a hard-feasible solution.",
        span=10)
    autosize_columns(ws)


def build_captions(wb: Workbook, merged: Dict[str, Any]) -> None:
    # Boxplot caption from FE = 1000, n_iter = 200 block
    ws = wb.create_sheet("7. Boxplot caption")
    r = 1
    r = big_title(ws, r,
        "Boxplot caption  |  boxplot_best_so_far_score_fe1000_n200.png",
        span=4)
    r += 1
    r = add_note(ws, r,
        "Figure (Boxplot).  Distribution of best-so-far total_score at FE = 1000 "
        "evaluator calls, n_iter = 200, NP = 20, 30 independent runs per method.  "
        "Lower is better.",
        span=4)
    r = section_title(ws, r,
        "Key statistics at FE = 1000, n_iter = 200, NP = 20", span=4)
    headers = ["method", "mean_best_so_far_total_score",
               "std_best_so_far_total_score", "feasible_rate_at_budget"]
    boxplot_rows: List[List[Any]] = []
    for method in ALL_METHODS_CANONICAL:
        row = next((row for row in merged["fe_budget"]
                    if row["method"] == method
                    and int(row["n_iter"]) == 200
                    and int(row["fe_budget"]) == 1000), None)
        if row is None:
            boxplot_rows.append([method, "-", "-", "-"])
        else:
            boxplot_rows.append([
                method,
                fnum(row["mean_best_so_far_total_score"]),
                fnum(row["std_best_so_far_total_score"]),
                fnum(row["feasible_rate_at_budget"]),
            ])
    num = "#,##0.00"
    r = write_table(ws, r, headers, boxplot_rows,
                    number_formats={1: num, 2: num, 3: "0.00%"})
    r += 1
    r = add_note(ws, r,
        "Source row: fe_budget_summary*.csv where n_iter = 200 and fe_budget = 1000.  "
        "Avoid 'N = 200' wording; use 'n_iter = 200'.",
        span=4)
    autosize_columns(ws)

    # Convergence caption from summary at n_iter = 200
    ws2 = wb.create_sheet("8. Convergence caption")
    r = 1
    r = big_title(ws2, r,
        "Convergence caption  |  combined_convergence_n200.png",
        span=4)
    r += 1
    r = add_note(ws2, r,
        "Figure (Convergence).  Best-so-far total_score vs outer iteration, "
        "n_iter = 200, NP = 20, mean of 30 runs.  Initial point "
        "(outer_iter = 0, label = greedy) is shared: total_score = 6,248,436.98 "
        "(= 5,989,692.00 cost + 258,744.98 hard).",
        span=4)
    r = section_title(ws2, r,
        "Final-iteration statistics at n_iter = 200, NP = 20", span=4)
    headers = ["method", "mean_total_score", "std", "rel. improvement vs greedy"]
    summary_n_by_key = {(r["method"], int(r["n_iter"])): r for r in merged["summary_n"]}
    greedy = 6248436.975833333
    conv_rows: List[List[Any]] = []
    for method in ALL_METHODS_CANONICAL:
        row = summary_n_by_key.get((method, 200))
        if row is None:
            conv_rows.append([method, "-", "-", "-"])
        else:
            mean_val = fnum(row["mean"])
            conv_rows.append([
                method,
                mean_val,
                fnum(row["std_dev"]),
                (mean_val / greedy) - 1.0 if mean_val else None,
            ])
    r = write_table(ws2, r, headers, conv_rows,
                    number_formats={1: num, 2: num, 3: "0.00%"})
    r += 1
    r = add_note(ws2, r,
        "Source: summary_by_method_n.csv (mean, std at n_iter = 200) + "
        "convergence_history*.csv (greedy reference).",
        span=4)
    autosize_columns(ws2)


def build_provenance(wb: Workbook, runs: List[Dict[str, Any]],
                     missing: List[str]) -> None:
    ws = wb.create_sheet("9. Provenance & extension")
    r = 1
    r = big_title(ws, r, "Per-(method, n_iter) provenance and how to extend", span=4)
    r += 1

    # Provenance: which run contributed which cell
    cells: Dict[Tuple[str, int], str] = {}
    for run in runs:
        rel = str(run["dir"].relative_to(ROOT))
        for row in run["summary_n"]:
            cells[(row["method"], int(row["n_iter"]))] = rel

    r = section_title(ws, r, "Cell provenance (which run contributed each cell)", span=4)
    rows: List[List[Any]] = []
    for method in ALL_METHODS_CANONICAL:
        for n in N_ITER_CANONICAL:
            rows.append([method, NP_VALUE, n,
                         cells.get((method, n), "-")])
    r = write_table(ws, r,
                    ["method", "NP", "n_iter", "source run directory"],
                    rows)
    r += 1

    r = section_title(ws, r, "How to extend to a full all-method comparison", span=4)
    steps = [
        ("Step 1", "Run api_benchmark_runner.py for the missing methods "
                   "(" + (", ".join(missing) if missing else "none -- already complete") + ") "
                   "with NP = 20, n_iter in {10, 20, 30, 50, 100, 200}, runs_per_n = 30, "
                   "on the same DB_SAMPLE input."),
        ("Step 2", "Output directory should be Doc/benchmark_outputs/paper_comparison_main_<N>/ "
                   "(e.g. _3), containing manifest.json and csv/ subfolder with the standard "
                   "summary_by_method_n.csv, summary_by_method.csv, fairness_by_method_n.csv, "
                   "fairness_by_method.csv, fe_budget_summary.csv, convergence_history.csv."),
        ("Step 3", "Re-run Doc/build_tables_xlsx.py.  The script auto-discovers every run "
                   "directory and rebuilds tables_report.xlsx.  Placeholders (-) become "
                   "real numbers automatically; sheet titles drop the 'NP=20 re-evaluation' "
                   "warning when all 8 methods are present."),
    ]
    r = write_table(ws, r, ["Step", "Action"], steps)
    autosize_columns(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    runs_dirs = discover_runs(BENCH_ROOT)
    if not runs_dirs:
        raise SystemExit(f"No benchmark runs found under {BENCH_ROOT}")
    if not INPUT_JSON.exists():
        raise SystemExit(f"example_input.json not found: {INPUT_JSON}")

    runs = [load_run(d) for d in runs_dirs]
    merged = merge_runs(runs)
    present = methods_present(merged)
    missing = methods_missing(present)
    instance = json.loads(INPUT_JSON.read_text(encoding="utf-8"))

    wb = Workbook()
    wb.remove(wb.active)

    build_scope(wb, present, missing)
    build_sources(wb, runs)
    build_instance(wb, instance["problemData"])
    build_fixed_iter(wb, merged, present, missing)
    build_fe_budget(wb, merged, present, missing)
    build_runtime(wb, merged, present, missing)
    build_hard_total(wb, merged, present, missing)
    build_captions(wb, merged)
    build_provenance(wb, runs, missing)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote   : {OUT_XLSX}")
    print(f"Size    : {OUT_XLSX.stat().st_size:,} bytes")
    print(f"Sheets  : {wb.sheetnames}")
    print(f"Runs    : {[str(r['dir'].relative_to(ROOT)) for r in runs]}")
    print(f"Methods : present = {present}  missing = {missing}")


if __name__ == "__main__":
    main()
