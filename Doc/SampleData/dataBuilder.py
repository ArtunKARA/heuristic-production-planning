#!/usr/bin/env python3
"""
Build example input from PostgreSQL using stocks/orders + vardiya templates.

Outputs a ProblemFrame-style JSON:
  { "problemData": {...}, "scenarioConfig": {...}, "state": {...} }

Notes:
- Uses mrp-api schema by default (quoted, because of hyphen).
- Reads products from stocks_and_orders.json and enriches from DB.
- Best-effort: missing DB data will fall back to placeholders.

runfile('.../dataBuilder.py', wdir='...', args='--user postgres --password dasdasbursa')
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse


try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except Exception:  # pragma: no cover - handled by runtime message
    psycopg2 = None  # type: ignore
    RealDictCursor = None  # type: ignore

try:
    import openpyxl
except Exception:  # pragma: no cover - optional
    openpyxl = None  # type: ignore


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_STOCKS = BASE_DIR / "stocks_and_orders.json"
DEFAULT_VARDIYA = BASE_DIR / "vardiya.json"
DEFAULT_DICT = BASE_DIR / "data_dictionary.xlsx"
DEFAULT_OUT = BASE_DIR / "example_input.json"
DEFAULT_JDBC = "jdbc:postgresql://localhost:5432/MRPFerkan"
DEFAULT_SCHEMA = "mrp-api"


def parse_jdbc_url(url: str) -> Tuple[str, int, str]:
    # Accepts:
    #   jdbc:postgresql://host:5432/db
    #   postgresql://host:5432/db
    #   host:5432/db
    if url is None:
        raise ValueError("Invalid JDBC url: None")
    raw = str(url).strip()
    # remove non-printable chars (hidden unicode / control chars)
    cleaned = "".join(ch for ch in raw if ch.isprintable())
    if cleaned:
        raw = cleaned.strip()

    if raw.startswith("jdbc:"):
        raw = raw[len("jdbc:") :]

    if not raw.startswith("postgresql://"):
        raw = "postgresql://" + raw.lstrip("/")

    parsed = urlparse(raw)
    host = parsed.hostname
    port = parsed.port or 5432
    db = (parsed.path or "").lstrip("/")

    if not host or not db:
        raise ValueError(f"Invalid JDBC url: {url!r}")

    return host, port, db


def load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def iso_week_range(week_id: str) -> Tuple[date, date]:
    # CW43_25 -> year=2025, week=43
    if week_id is None:
        raise ValueError("Invalid week id: None")
    week_id = str(week_id).strip()
    m = re.match(r"^CW(\d{1,2})_(\d{2,4})$", week_id)
    if not m:
        raise ValueError(f"Invalid week id: {week_id!r}")
    week = int(m.group(1))
    year = int(m.group(2))
    if year < 100:
        year += 2000

    # clamp invalid ISO weeks if needed
    max_week = date(year, 12, 28).isocalendar()[1]
    if week > max_week:
        week = max_week

    start = date.fromisocalendar(year, week, 1)
    end = date.fromisocalendar(year, week, 7)
    return start, end


def bucket_sort_key(week_id: str) -> Tuple[int, int]:
    if week_id is None:
        return (9999, 999)
    week_id = str(week_id).strip()
    m = re.match(r"^CW(\d{1,2})_(\d{2,4})$", week_id)
    if not m:
        return (9999, 999)
    week = int(m.group(1))
    year = int(m.group(2))
    if year < 100:
        year += 2000
    return (year, week)


def build_time_buckets(orders: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    weeks: List[str] = []
    for og in orders:
        for it in og.get("orders", []):
            wk = it.get("week") or it.get("time_bucket_id")
            if wk is not None:
                wk = str(wk).strip()
            if wk:
                weeks.append(wk)

    unique = sorted(set(weeks), key=bucket_sort_key)
    buckets: List[Dict[str, Any]] = []
    for idx, wk in enumerate(unique):
        start, end = iso_week_range(wk)
        buckets.append({"id": wk, "index": idx, "start_date": start, "end_date": end})
    return buckets


def normalize_orders(orders: List[Dict[str, Any]], bucket_by_id: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    normalized = []
    for og in orders:
        code = og.get("product_code")
        by_week: Dict[str, float] = defaultdict(float)
        for it in og.get("orders", []):
            wk = it.get("week") or it.get("time_bucket_id")
            if wk is not None:
                wk = str(wk).strip()
            if not wk:
                continue
            by_week[wk] += float(it.get("qty", 0) or 0)
        items = []
        for i, (wk, qty) in enumerate(sorted(by_week.items(), key=lambda x: bucket_sort_key(x[0]))):
            tb = bucket_by_id.get(wk)
            due = None
            if tb:
                due = datetime.combine(tb["end_date"], datetime.max.time()).replace(microsecond=0)
            items.append(
                {
                    "order_id": f"{code}_{wk}_{i+1}",
                    "time_bucket_id": wk,
                    "due_date": due,
                    "qty": qty,
                }
            )
        normalized.append({"product_code": code, "orders": items})
    return normalized


def parse_hhmm(val: str) -> int:
    h, m = val.split(":")
    return int(h) * 60 + int(m)


def shift_template_hours(segments: List[Dict[str, Any]]) -> float:
    total = 0.0
    for seg in segments:
        start = parse_hhmm(seg["start"])
        end = parse_hhmm(seg["end"])
        if end <= start:
            end += 24 * 60
        total += (end - start) / 60.0
    return total


def build_shift_templates(vardiya: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    templates = []
    base_code = None
    for idx, tpl in enumerate(vardiya.get("shift_templates", [])):
        code = tpl.get("id") or tpl.get("code")
        if base_code is None:
            base_code = code
        segments = []
        for seg in tpl.get("segments", []):
            end_time = seg.get("end")
            # normalize 24:00 -> 00:00 to avoid invalid hour in downstream datetime
            if end_time == "24:00":
                end_time = "00:00"
            constraints = []
            if seg.get("mold_change_allowed") is False:
                constraints.append("NO_MOLD_CHANGE_AT_NIGHT")
            segments.append(
                {
                    "code": seg.get("code"),
                    "start": seg.get("start"),
                    "end": end_time,
                    "constraints": constraints,
                }
            )
        templates.append({"code": code, "name": tpl.get("name"), "segments": segments})
    return templates, base_code


def build_work_calendar(time_buckets: List[Dict[str, Any]], base_code: str) -> List[Dict[str, Any]]:
    if not time_buckets:
        return []
    entries = []
    min_date = min(tb["start_date"] for tb in time_buckets)
    max_date = max(tb["end_date"] for tb in time_buckets)
    d = min_date
    while d <= max_date:
        entries.append({"date": d, "shift_templates_code": base_code, "holiday": False})
        d += timedelta(days=1)
    return entries


def json_default(obj: Any) -> Any:
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    raise TypeError(f"Type not serializable: {type(obj)}")


def db_connect(jdbc_url: str, user: str, password: Optional[str]) -> Any:
    if psycopg2 is None:
        raise RuntimeError("psycopg2 is not installed. Install with: pip install psycopg2-binary")
    host, port, db = parse_jdbc_url(jdbc_url)
    return psycopg2.connect(host=host, port=port, dbname=db, user=user, password=password)


def q(schema: str, table: str) -> str:
    return f"\"{schema}\".{table}"


def fetch_all(cur, sql: str, params: Optional[Tuple[Any, ...]] = None) -> List[Dict[str, Any]]:
    cur.execute(sql, params or ())
    return list(cur.fetchall())


def fetch_materials_by_code(cur, schema: str, codes: List[str]) -> List[Dict[str, Any]]:
    if not codes:
        return []
    sql = f"""
        SELECT m.id, m.code, m.name, m.usage_unit_id,
               m.material_bom_type_id, m.material_type_id,
               u.code AS unit_code, u.name AS unit_name
        FROM {q(schema, 'material')} m
        LEFT JOIN {q(schema, 'mesure')} u ON u.id = m.usage_unit_id
        WHERE m.code = ANY(%s)
    """
    return fetch_all(cur, sql, (codes,))


def fetch_materials_by_id(cur, schema: str, ids: List[int]) -> List[Dict[str, Any]]:
    if not ids:
        return []
    sql = f"""
        SELECT m.id, m.code, m.name, m.usage_unit_id,
               m.material_bom_type_id, m.material_type_id,
               u.code AS unit_code, u.name AS unit_name
        FROM {q(schema, 'material')} m
        LEFT JOIN {q(schema, 'mesure')} u ON u.id = m.usage_unit_id
        WHERE m.id = ANY(%s)
    """
    return fetch_all(cur, sql, (ids,))


def fetch_bom_revisions(cur, schema: str, material_ids: List[int]) -> Dict[int, int]:
    if not material_ids:
        return {}
    sql = f"""
        SELECT DISTINCT ON (material_id) id, material_id
        FROM {q(schema, 'bom_revision')}
        WHERE material_id = ANY(%s) AND bom_use = true
        ORDER BY material_id, id DESC
    """
    rows = fetch_all(cur, sql, (material_ids,))
    return {r["material_id"]: r["id"] for r in rows}


def fetch_process_receipt_revisions(cur, schema: str, material_ids: List[int], bom_rev_ids: List[int]) -> Dict[int, Dict[str, Any]]:
    revs: Dict[int, Dict[str, Any]] = {}
    if material_ids:
        sql = f"""
            SELECT DISTINCT ON (material_id) id, material_id, bom_revision_id, process_bom_id
            FROM {q(schema, 'process_receipt_revision')}
            WHERE material_id = ANY(%s) AND process_receipt_use = true
            ORDER BY material_id, id DESC
        """
        rows = fetch_all(cur, sql, (material_ids,))
        for r in rows:
            revs[r["material_id"]] = r

    missing_bom = [bid for bid in bom_rev_ids if bid and bid not in {r.get("bom_revision_id") for r in revs.values()}]
    if missing_bom:
        sql = f"""
            SELECT DISTINCT ON (bom_revision_id) id, material_id, bom_revision_id, process_bom_id
            FROM {q(schema, 'process_receipt_revision')}
            WHERE bom_revision_id = ANY(%s) AND process_receipt_use = true
            ORDER BY bom_revision_id, id DESC
        """
        rows = fetch_all(cur, sql, (missing_bom,))
        for r in rows:
            if r.get("material_id") and r["material_id"] not in revs:
                revs[r["material_id"]] = r
    return revs


def fetch_process_receipts(cur, schema: str, rev_ids: List[int]) -> List[Dict[str, Any]]:
    if not rev_ids:
        return []
    sql = f"""
        SELECT id, process_receipt_revision_id, process_definition_id,
               operation_code, base_quantity, base_quantity_type_id,
               setup_time, setup_time_type_id, cycle_time, cycle_time_type_id,
               waiting_time, waiting_time_type_id, adding_material_id
        FROM {q(schema, 'process_receipt')}
        WHERE process_receipt_revision_id = ANY(%s)
        ORDER BY process_receipt_revision_id, operation_code NULLS LAST, id
    """
    return fetch_all(cur, sql, (rev_ids,))


def fetch_process_definitions(cur, schema: str, ids: List[int]) -> Dict[int, Dict[str, Any]]:
    if not ids:
        return {}
    sql = f"""
        SELECT id, code, name
        FROM {q(schema, 'process_definition')}
        WHERE id = ANY(%s)
    """
    rows = fetch_all(cur, sql, (ids,))
    return {r["id"]: r for r in rows}


def fetch_process_bom(cur, schema: str, bom_rev_ids: List[int]) -> List[Dict[str, Any]]:
    if not bom_rev_ids:
        return []
    sql = f"""
        SELECT id, bom_revision_id, material_id, quantity, quantity_type_id
        FROM {q(schema, 'process_bom')}
        WHERE bom_revision_id = ANY(%s)
    """
    return fetch_all(cur, sql, (bom_rev_ids,))


def fetch_mesures(cur, schema: str, ids: List[int]) -> Dict[int, Dict[str, Any]]:
    if not ids:
        return {}
    sql = f"""
        SELECT id, code, name
        FROM {q(schema, 'mesure')}
        WHERE id = ANY(%s)
    """
    rows = fetch_all(cur, sql, (ids,))
    return {r["id"]: r for r in rows}


def fetch_machines(
    cur,
    schema: str,
    process_ids: Optional[List[int]] = None,
    machine_group_ids: Optional[List[int]] = None,
) -> List[Dict[str, Any]]:
    rows: Dict[int, Dict[str, Any]] = {}

    if process_ids:
        sql = f"""
            SELECT id, name, process_id, machine_group_id, machine_active
            FROM {q(schema, 'machine')}
            WHERE process_id = ANY(%s)
        """
        for r in fetch_all(cur, sql, (process_ids,)):
            rows[r["id"]] = r

    if machine_group_ids:
        sql = f"""
            SELECT id, name, process_id, machine_group_id, machine_active
            FROM {q(schema, 'machine')}
            WHERE machine_group_id = ANY(%s)
        """
        for r in fetch_all(cur, sql, (machine_group_ids,)):
            rows[r["id"]] = r

    return list(rows.values())


def fetch_machine_groups(cur, schema: str, ids: List[int]) -> List[Dict[str, Any]]:
    if not ids:
        return []
    sql = f"""
        SELECT id, name, code, process_id
        FROM {q(schema, 'machine_group')}
        WHERE id = ANY(%s)
    """
    return fetch_all(cur, sql, (ids,))


def fetch_process_molds(cur, schema: str, material_ids: List[int]) -> List[Dict[str, Any]]:
    if not material_ids:
        return []
    sql = f"""
        SELECT material_id, mold_id
        FROM {q(schema, 'process_mold')}
        WHERE material_id = ANY(%s)
    """
    return fetch_all(cur, sql, (material_ids,))


def fetch_molds(cur, schema: str, mold_ids: List[int]) -> List[Dict[str, Any]]:
    if not mold_ids:
        return []
    sql = f"""
        SELECT id, name, mold_type_id, eye_count
        FROM {q(schema, 'mold')}
        WHERE id = ANY(%s)
    """
    return fetch_all(cur, sql, (mold_ids,))


def fetch_mold_type_machine_group(cur, schema: str, mold_type_ids: List[int]) -> List[Dict[str, Any]]:
    if not mold_type_ids:
        return []
    sql = f"""
        SELECT mold_type_id, machine_group_id
        FROM {q(schema, 'mold_type_machine_group')}
        WHERE mold_type_id = ANY(%s)
    """
    return fetch_all(cur, sql, (mold_type_ids,))


def unit_label(measure: Optional[Dict[str, Any]]) -> Optional[str]:
    if not measure:
        return None
    return measure.get("code") or measure.get("name")


def convert_time(value: Any, unit_code: Optional[str], target: str) -> float:
    if value is None:
        return 0.0
    try:
        v = float(value)
    except Exception:
        return 0.0
    if not unit_code:
        return v
    u = unit_code.strip().lower()
    # basic mapping
    seconds = v
    if u in {"sn", "sec", "s", "second", "seconds"}:
        seconds = v
    elif u in {"dk", "min", "minute", "minutes"}:
        seconds = v * 60.0
    elif u in {"saat", "hour", "hours", "h", "hr"}:
        seconds = v * 3600.0
    elif u in {"gun", "day", "days"}:
        seconds = v * 86400.0
    if target == "sec":
        return seconds
    if target == "min":
        return seconds / 60.0
    if target == "hour":
        return seconds / 3600.0
    return v


def check_dictionary(path: Path, required_tables: List[str]) -> List[str]:
    if openpyxl is None or not path.exists():
        return []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "tables" not in wb.sheetnames:
        return []
    ws = wb["tables"]
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 3 and row[2]:
            names.add(str(row[2]))
    missing = [t for t in required_tables if t not in names]
    return missing


def main() -> int:
    parser = argparse.ArgumentParser(description="Build example input from DB.")
    parser.add_argument("--stocks", default=str(DEFAULT_STOCKS))
    parser.add_argument("--vardiya", default=str(DEFAULT_VARDIYA))
    parser.add_argument("--dictionary", default=str(DEFAULT_DICT))
    parser.add_argument("--output", default=str(DEFAULT_OUT))
    parser.add_argument("--jdbc", default=os.getenv("JDBC_URL", DEFAULT_JDBC))
    parser.add_argument("--schema", default=os.getenv("PGSCHEMA", DEFAULT_SCHEMA))
    parser.add_argument("--user", default=os.getenv("PGUSER", "postgres"))
    parser.add_argument("--password", default=os.getenv("PGPASSWORD"))
    args = parser.parse_args()

    stocks_path = Path(args.stocks)
    vardiya_path = Path(args.vardiya)
    dict_path = Path(args.dictionary)
    out_path = Path(args.output)

    if not stocks_path.exists():
        print(f"Missing stocks file: {stocks_path}", file=sys.stderr)
        return 1
    if not vardiya_path.exists():
        print(f"Missing vardiya file: {vardiya_path}", file=sys.stderr)
        return 1

    missing = check_dictionary(
        dict_path,
        [
            "material",
            "bom_revision",
            "process_receipt_revision",
            "process_receipt",
            "process_definition",
            "process_bom",
            "machine",
            "process_mold",
            "mold",
            "mold_type_machine_group",
            "mesure",
        ],
    )
    if missing:
        print(f"Warning: missing tables in data_dictionary.xlsx: {', '.join(missing)}", file=sys.stderr)

    stocks_orders = load_json(stocks_path)
    vardiya = load_json(vardiya_path)

    orders_raw = stocks_orders.get("orders", [])
    stocks_raw = stocks_orders.get("stocks", [])

    shift_templates, base_shift_code = build_shift_templates(vardiya)
    if not base_shift_code:
        base_shift_code = "S3"

    # deduce product codes from stocks/orders
    product_codes = sorted(
        {it.get("product_code") for it in (orders_raw + stocks_raw) if it.get("product_code")}
    )

    if not product_codes:
        print("No product codes found in stocks/orders.", file=sys.stderr)
        return 1

    # connect DB
    try:
        conn = db_connect(args.jdbc, args.user, args.password)
    except Exception as exc:
        print(f"DB connection failed: {exc}", file=sys.stderr)
        return 2

    with conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            materials = fetch_materials_by_code(cur, args.schema, product_codes)
            material_by_code = {m["code"]: m for m in materials}
            found_codes = set(material_by_code.keys())

            missing_codes = [c for c in product_codes if c not in found_codes]
            if missing_codes:
                print(
                    "Warning: product codes not found in DB (skipped in output): "
                    + ", ".join(missing_codes),
                    file=sys.stderr,
                )

            # filter orders/stocks to only DB-found products
            orders_raw = [og for og in orders_raw if og.get("product_code") in found_codes]
            stocks_raw = [st for st in stocks_raw if st.get("product_code") in found_codes]
            product_codes = sorted(found_codes)

            if not product_codes:
                print("No product codes found in DB after filtering.", file=sys.stderr)
                return 1

            product_ids = [m["id"] for m in materials if m.get("code") in found_codes]

            bom_rev_by_mat = fetch_bom_revisions(cur, args.schema, product_ids)
            bom_rev_ids = list(bom_rev_by_mat.values())

            pr_rev_by_mat = fetch_process_receipt_revisions(cur, args.schema, product_ids, bom_rev_ids)
            pr_rev_ids = [r["id"] for r in pr_rev_by_mat.values()]
            process_receipts = fetch_process_receipts(cur, args.schema, pr_rev_ids)

            process_ids = {r["process_definition_id"] for r in process_receipts if r.get("process_definition_id")}
            process_defs = fetch_process_definitions(cur, args.schema, sorted(process_ids))

            process_bom = fetch_process_bom(cur, args.schema, bom_rev_ids)

            # measures
            measure_ids = set()
            for r in process_receipts:
                for key in ("base_quantity_type_id", "setup_time_type_id", "cycle_time_type_id", "waiting_time_type_id"):
                    if r.get(key):
                        measure_ids.add(r[key])
            for r in process_bom:
                if r.get("quantity_type_id"):
                    measure_ids.add(r["quantity_type_id"])
            measures = fetch_mesures(cur, args.schema, list(measure_ids))

            # process steps by product
            steps_by_product_id: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
            rev_to_material = {rev.get("id"): mat_id for mat_id, rev in pr_rev_by_mat.items()}
            for r in process_receipts:
                mat_id = rev_to_material.get(r.get("process_receipt_revision_id"))
                if mat_id:
                    steps_by_product_id[mat_id].append(r)

            # build input materials from BOM
            inputs_by_bom_rev: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
            for r in process_bom:
                inputs_by_bom_rev[r["bom_revision_id"]].append(r)

            # collect input material ids
            input_material_ids = set()
            for rows in inputs_by_bom_rev.values():
                for r in rows:
                    if r.get("material_id"):
                        input_material_ids.add(r["material_id"])
            for r in process_receipts:
                if r.get("adding_material_id"):
                    input_material_ids.add(r["adding_material_id"])

            extra_materials = fetch_materials_by_id(cur, args.schema, list(input_material_ids))
            material_by_id = {m["id"]: m for m in materials + extra_materials}

            # molds -> machine groups
            process_molds = fetch_process_molds(cur, args.schema, product_ids)
            mold_ids = sorted({r["mold_id"] for r in process_molds if r.get("mold_id")})
            molds_raw = fetch_molds(cur, args.schema, mold_ids)

            mold_type_ids = sorted({m["mold_type_id"] for m in molds_raw if m.get("mold_type_id")})
            mtmg_rows = fetch_mold_type_machine_group(cur, args.schema, mold_type_ids)
            mtmg_by_mold_type: Dict[int, List[int]] = defaultdict(list)
            for r in mtmg_rows:
                mtmg_by_mold_type[r["mold_type_id"]].append(r["machine_group_id"])

            machine_group_ids = sorted({r["machine_group_id"] for r in mtmg_rows if r.get("machine_group_id")})
            machine_groups = fetch_machine_groups(cur, args.schema, machine_group_ids)
            machine_group_by_id = {mg["id"]: mg for mg in machine_groups}

            # include machine-group processes in process definitions
            process_ids.update({mg["process_id"] for mg in machine_groups if mg.get("process_id")})

            # machines (by process and by machine_group)
            machines = fetch_machines(
                cur,
                args.schema,
                process_ids=sorted(process_ids),
                machine_group_ids=machine_group_ids,
            )
            machines = [m for m in machines if m.get("machine_active") is not False]

            # include machine processes too
            process_ids.update({m["process_id"] for m in machines if m.get("process_id")})
            process_defs = fetch_process_definitions(cur, args.schema, sorted(process_ids))

    # build process steps
    products_out = []
    processes_out: Dict[str, Dict[str, Any]] = {}

    for code in product_codes:
        mat = material_by_code.get(code)
        if not mat:
            continue

        mat_id = mat["id"]
        base_unit = mat.get("unit_code") or mat.get("unit_name") or "ADET"

        steps_rows = steps_by_product_id.get(mat_id, [])
        steps_rows = sorted(steps_rows, key=lambda x: (x.get("operation_code") is None, x.get("operation_code") or 0, x.get("id") or 0))

        process_steps = []
        for idx, r in enumerate(steps_rows):
            proc = process_defs.get(r.get("process_definition_id")) or {}
            proc_code = proc.get("code") or f"P{r.get('process_definition_id')}"
            proc_name = proc.get("name") or proc_code

            base_qty_unit = unit_label(measures.get(r.get("base_quantity_type_id")))
            setup_unit = unit_label(measures.get(r.get("setup_time_type_id")))
            cycle_unit = unit_label(measures.get(r.get("cycle_time_type_id")))
            wait_unit = unit_label(measures.get(r.get("waiting_time_type_id"))) or "DAY"

            base_qty = float(r.get("base_quantity") or 1)
            setup_min = convert_time(r.get("setup_time"), setup_unit, "min")
            cycle_sec = convert_time(r.get("cycle_time"), cycle_unit, "sec")
            wait_time = float(r.get("waiting_time") or 0)

            step_no = int(r.get("operation_code") or (idx + 1) * 10)

            process_steps.append(
                {
                    "process_code": proc_code,
                    "step_no": step_no,
                    "name": proc_name,
                    "output_material": code,
                    "yield_factor": 1.0,
                    "base_qty": base_qty,
                    "base_qty_type": base_qty_unit or "ADET",
                    "setup_time_min": setup_min,
                    "cycle_time_sec": cycle_sec,
                    "wait_time": wait_time,
                    "wait_unit": wait_unit,
                    "inputs": [],
                }
            )

            processes_out[proc_code] = {"code": proc_code, "name": proc_name, "constraints": []}

        # attach process_bom inputs to first step
        bom_rev = bom_rev_by_mat.get(mat_id)
        if process_steps and bom_rev:
            inputs = []
            for inp in inputs_by_bom_rev.get(bom_rev, []):
                mat_info = material_by_id.get(inp.get("material_id"))
                if not mat_info:
                    continue
                qty_unit = unit_label(measures.get(inp.get("quantity_type_id"))) or base_unit
                try:
                    qty = float(inp.get("quantity") or 0)
                except Exception:
                    qty = 0.0
                qty_per = qty / float(process_steps[0]["base_qty"] or 1)
                inputs.append(
                    {
                        "material_code": mat_info.get("code"),
                        "qty_per_output_unit": qty_per,
                        "qty_unit": qty_unit,
                        "scrap_factor": 0.0,
                    }
                )
            process_steps[0]["inputs"] = inputs

        if process_steps:
            products_out.append(
                {
                    "code": mat.get("code"),
                    "name": mat.get("name") or mat.get("code"),
                    "base_unit": base_unit,
                    "process_data": process_steps,
                }
            )

    # if we dropped products due to missing process_data, filter orders/stocks/product_codes
    product_codes = sorted({p["code"] for p in products_out})
    orders_raw = [og for og in orders_raw if og.get("product_code") in set(product_codes)]
    stocks_raw = [st for st in stocks_raw if st.get("product_code") in set(product_codes)]

    # rebuild time buckets based on filtered orders
    time_buckets = build_time_buckets(orders_raw)
    bucket_by_id = {tb["id"]: tb for tb in time_buckets}
    work_calendar = build_work_calendar(time_buckets, base_shift_code)

    # materials list (product + inputs)
    material_list = []
    seen = set()
    for m in (list(material_by_code.values()) + list(material_by_id.values())):
        code = m.get("code")
        if not code or code in seen:
            continue
        seen.add(code)
        material_list.append(
            {
                "code": code,
                "name": m.get("name") or code,
                "base_unit": m.get("unit_code") or m.get("unit_name") or "ADET",
            }
        )

    # build machines
    process_def_by_id = {v["id"]: v for v in (process_defs.values())} if process_defs else {}
    process_def_by_code = {v.get("code"): v for v in (process_defs.values()) if v.get("code")} if process_defs else {}
    machines_out = []
    for m in machines:
        proc = process_def_by_id.get(m.get("process_id")) or {}
        proc_code = proc.get("code")
        if not proc_code:
            continue
        machines_out.append(
            {
                "id": m.get("id"),
                "name": m.get("name") or f"Machine {m.get('id')}",
                "process_code": proc_code,
            }
        )

    # capacity by bucket (same for each machine)
    hours_per_day = 24.0
    if shift_templates:
        hours_per_day = shift_template_hours(shift_templates[0]["segments"])
    cap_by_bucket = {}
    for tb in time_buckets:
        days = (tb["end_date"] - tb["start_date"]).days + 1
        cap_by_bucket[tb["id"]] = round(hours_per_day * days, 2)
    for m in machines_out:
        m["capacity_by_bucket"] = dict(cap_by_bucket)

    # molds and compatibility
    molds_out = []
    machine_by_group: Dict[int, List[int]] = defaultdict(list)
    for m in machines:
        mgid = m.get("machine_group_id")
        if mgid:
            machine_by_group[mgid].append(m.get("id"))

    molds_by_id = {m["id"]: m for m in molds_raw}
    product_molds_map: Dict[Tuple[str, str], set] = defaultdict(set)
    machine_process_code = {m["id"]: m.get("process_code") for m in machines_out}

    # map product -> process code (use last step)
    product_last_process: Dict[str, Optional[str]] = {}
    for p in products_out:
        steps = p.get("process_data", [])
        if steps:
            steps_sorted = sorted(steps, key=lambda x: x.get("step_no", 0))
            product_last_process[p["code"]] = steps_sorted[-1].get("process_code")
        else:
            product_last_process[p["code"]] = None

    for rel in process_molds:
        mat_id = rel.get("material_id")
        mold_id = rel.get("mold_id")
        mat = material_by_id.get(mat_id)
        mold = molds_by_id.get(mold_id)
        if not mat or not mold:
            continue
        prod_code = mat.get("code")
        mold_code = str(mold_id)
        mtid = mold.get("mold_type_id")
        allowed_groups = mtmg_by_mold_type.get(mtid, [])
        candidate_proc_codes: set = set()
        for gid in allowed_groups:
            mg = machine_group_by_id.get(gid, {})
            pid = mg.get("process_id")
            if pid:
                proc = process_defs.get(pid) or {}
                if proc.get("code"):
                    candidate_proc_codes.add(proc["code"])

        prod_proc = product_last_process.get(prod_code)
        if prod_proc in candidate_proc_codes:
            proc_code = prod_proc
        elif len(candidate_proc_codes) == 1:
            proc_code = next(iter(candidate_proc_codes))
        else:
            proc_code = prod_proc or (next(iter(candidate_proc_codes)) if candidate_proc_codes else None)

        compat_machines: List[int] = []
        for gid in allowed_groups:
            compat_machines.extend(machine_by_group.get(gid, []))
        if proc_code:
            compat_machines = [mid for mid in compat_machines if machine_process_code.get(mid) == proc_code]
        if not compat_machines:
            # fallback: all machines of the same process
            if proc_code:
                compat_machines = [m["id"] for m in machines_out if m.get("process_code") == proc_code]
            else:
                compat_machines = [m["id"] for m in machines_out]

        if not proc_code:
            continue

        product_molds_map[(prod_code, proc_code)].add(mold_code)
        if proc_code not in processes_out:
            pdef = process_def_by_code.get(proc_code, {})
            processes_out[proc_code] = {"code": proc_code, "name": pdef.get("name") or proc_code, "constraints": []}
        molds_out.append(
            {
                "code": mold_code,
                "name": mold.get("name") or f"Mold {mold_id}",
                "process_code": proc_code,
                "eye": mold.get("eye_count"),
                "compatible_machines_id": sorted(set(compat_machines)),
                "supported_products_id": [prod_code],
            }
        )

    # processes constraints
    process_constraints: Dict[str, set] = defaultdict(set)
    for m in machines_out:
        process_constraints[m["process_code"]].add("machine")
    for mold in molds_out:
        if mold.get("process_code"):
            process_constraints[mold["process_code"]].add("mold")
    for code, proc in processes_out.items():
        proc["constraints"] = sorted(process_constraints.get(code, set()))

    processes_list = list(processes_out.values())

    # compatibility lists
    machine_mold_pairs = []
    for mold in molds_out:
        proc_code = mold.get("process_code")
        for mid in mold.get("compatible_machines_id") or []:
            machine_mold_pairs.append(
                {"machine_id": mid, "mold_code": mold["code"], "process_code": proc_code}
            )

    product_molds = []
    for (pcode, proc_code), molds in product_molds_map.items():
        product_molds.append({"product_code": pcode, "process_code": proc_code, "allowed_molds": sorted(molds)})

    frame = {
        "problemData": {
            "problem_meta": {
                "problem_code": "DB_SAMPLE",
                "horizon_type": "Week",
                "base_shift_templates_code": base_shift_code,
            },
            "time_buckets": time_buckets,
            "orders": normalize_orders(orders_raw, bucket_by_id),
            "stocks": stocks_raw,
            "materials": material_list,
            "products": products_out,
            "processes": processes_list,
            "resources": {"machines": machines_out, "molds": molds_out},
            "shift_templates": shift_templates,
            "work_calendar": work_calendar,
            "compatibility": {
                "machine_mold_pairs": machine_mold_pairs,
                "product_molds": product_molds,
            },
        },
        "scenarioConfig": {
            "meta": {"name": "Base_Scenario"},
            "weights": {
                "w_mold_change": 10.0,
                "w_night_mold_change": 100.0,
                "w_inventory": 1.0,
                "w_machine_count": 5.0,
            },
            "toggles": {
                "HARD_DUE_DATE_FULFILLMENT": True,
                "HARD_RESOURCE_ROLE_ASSIGNED": True,
                "HARD_TIME_BUCKET_VALID": True,
                "HARD_NO_HOLIDAY_WORK": True,
                "HARD_COMPAT_MACHINE_MOLD_PROCESS": True,
                "HARD_COMPAT_PRODUCT_MOLD": True,
                "HARD_CAPACITY_BUCKET": True,
                "HARD_CAPACITY_SEGMENT": True,
                "SOFT_MOLD_CHANGE_MINIMIZE": True,
                "SOFT_NIGHT_MOLD_CHANGE": True,
                "SOFT_INVENTORY_LOW": True,
                "SOFT_MACHINE_COUNT_LOW": True,
            },
        },
        "state": {"meta": {"iteration": 0}, "lots": []},
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(frame, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    print(f"Wrote example input to: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
